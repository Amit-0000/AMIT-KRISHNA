"""Builds performance/VoiceGuard_Performance_Improvement_Report.xlsx.

Uses performance/VoiceGuard_Baseline_Load_Test_Report.xlsx (built by
build_excel.py) as the formatting/layout template: same palette, header
style, banding, conditional-formatting rules, title/subtitle pattern, and
gridlines-off sheet styling. Data comes from:
  performance/before/baseline_results.json        (pre-fix baseline)
  performance/baseline_results.json               (post-fix, current)
  performance/performance_before_vs_after.json    (diffed comparison)

Does NOT read or write the original baseline workbook — output goes to a new
file in the same directory, per the request.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).parent
BEFORE = json.loads((ROOT / "before" / "baseline_results.json").read_text(encoding="utf-8"))
AFTER = json.loads((ROOT / "baseline_results.json").read_text(encoding="utf-8"))
COMPARISON = json.loads((ROOT / "performance_before_vs_after.json").read_text(encoding="utf-8"))
OUT_XLSX = ROOT / "VoiceGuard_Performance_Improvement_Report.xlsx"
ORIGINAL_XLSX = ROOT / "VoiceGuard_Baseline_Load_Test_Report.xlsx"

assert ORIGINAL_XLSX.exists(), "original baseline workbook must exist as the formatting template"

# ─── Palette / styles — identical to build_excel.py's template ────────────
NAVY = "1F2937"
ACCENT = "2563EB"
LIGHT_BAND = "F3F4F6"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
GREEN_FONT = "1E7B34"
YELLOW = "FFEB9C"
YELLOW_FONT = "9C6500"
RED = "FFC7CE"
RED_FONT = "9C0006"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=16)
SUBTITLE_FONT = Font(italic=True, color="6B7280", size=10)
SECTION_FONT = Font(bold=True, color=ACCENT, size=12)
LABEL_FONT = Font(bold=True, color=NAVY)
BAND_FILL = PatternFill("solid", fgColor=LIGHT_BAND)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
GREEN_FILL = PatternFill("solid", fgColor=GREEN)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)
RED_FILL = PatternFill("solid", fgColor=RED)


def style_header_row(ws: Worksheet, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def band_rows(ws: Worksheet, start_row: int, end_row: int, n_cols: int):
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = BAND_FILL
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = BORDER


def autosize(ws: Worksheet, widths: dict[str, int] | None = None, min_width=10, max_width=70):
    widths = widths or {}
    col_max = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            length = len(str(cell.value))
            col_max[col] = max(col_max.get(col, 0), length)
    for col, length in col_max.items():
        w = widths.get(col, max(min_width, min(max_width, length + 3)))
        ws.column_dimensions[col].width = w


def write_title(ws: Worksheet, title: str, subtitle: str = ""):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT


def verdict_cell(ws: Worksheet, cell: str, good: bool, text: str, size: int = 14):
    c = ws[cell]
    c.value = text
    c.font = Font(bold=True, size=size, color=GREEN_FONT if good else RED_FONT)
    c.fill = GREEN_FILL if good else RED_FILL
    c.alignment = CENTER


def response_time_conditional_format(ws: Worksheet, cell_range: str):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["200"], fill=GREEN_FILL, font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["200", "500"], fill=YELLOW_FILL, font=Font(color=YELLOW_FONT)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["500"], fill=RED_FILL, font=Font(color=RED_FONT, bold=True)),
    )


def pct_conditional_format(ws: Worksheet, cell_range: str, good_above: float):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=[str(good_above)], fill=GREEN_FILL, font=Font(color=GREEN_FONT, bold=True)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=[str(good_above)], fill=RED_FILL, font=Font(color=RED_FONT)),
    )


b_overall = BEFORE["overall_performance"]
a_overall = AFTER["overall_performance"]
overall_cmp = COMPARISON["overall_comparison"]
endpoint_cmp = COMPARISON["endpoint_comparison"]
resource_cmp = COMPARISON["resource_comparison"]
status_cmp = COMPARISON["http_status_comparison"]


def cmp_metric(name: str) -> dict:
    return next(x for x in overall_cmp if x["metric"] == name)


wb = Workbook()

# ═════════════════════════════════════════════════════════════════════════
# Sheet 1: Executive Summary
# ═════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Executive Summary"
write_title(ws, "VoiceGuard Performance Improvement Report", "Comparison of Baseline vs Optimized System")

verdict_cell(ws, "B4", True, "IMPROVED", size=16)
ws["A4"] = "Overall Result"
ws["A4"].font = LABEL_FONT
ws.row_dimensions[4].height = 24

rows = [
    ("Benchmark", "100 Virtual Users, 60 Seconds"),
    ("Environment", "Same Docker Stack, Same Hardware, Same k6 Configuration"),
    ("Primary Bottleneck", "bcrypt.checkpw()/hashpw() executed synchronously inside async authentication "
        "routes, blocking the asyncio event loop"),
    ("Fix Applied", "Password hashing/verification moved to worker threads via asyncio.to_thread() — "
        "bcrypt algorithm, 12 rounds, and hash format all unchanged"),
    ("Secondary Fixes", "PostgreSQL connection pool 5/20 -> 10/50; Redis connection pool 10 -> 50 "
        "(configuration-only changes, exposed only once the event-loop block was removed)"),
    ("Test Harness Fix", "k6 cookie-persistence issue identified and fixed by explicitly capturing and "
        "replaying the session cookie; benchmark configuration otherwise unchanged"),
]
r = 6
for label, value in rows:
    ws.cell(row=r, column=1, value=label).font = LABEL_FONT
    ws.cell(row=r, column=2, value=value).alignment = LEFT
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2).border = BORDER
    ws.row_dimensions[r].height = 32
    if r % 2 == 0:
        ws.cell(row=r, column=1).fill = BAND_FILL
        ws.cell(row=r, column=2).fill = BAND_FILL
    r += 1

headline_start = r + 2
ws.cell(row=headline_start, column=1, value="Headline Results").font = SECTION_FONT
hh = headline_start + 1
for c, h in enumerate(["Metric", "Before", "After", "Improvement"], start=1):
    ws.cell(row=hh, column=c, value=h)
style_header_row(ws, hh, 4)
headline_rows = [
    ("Success Rate", "1.25%", "86.19% (100% excl. expected 409s)", "Significant Improvement"),
    ("Average Response Time", "1736.9 ms", "176.7 ms", "-89.8%"),
    ("P99 Latency", "32340.7 ms", "1628.8 ms", "-95.0%"),
    ("Login Success", "15.5% (29.1 s avg)", "100% (323.5 ms avg)", "Fully Resolved"),
    ("401 Errors", "1424", "0", "Eliminated"),
    ("500 Errors", "71", "0", "Eliminated"),
]
rr = hh
for row in headline_rows:
    rr += 1
    for c, val in enumerate(row, start=1):
        ws.cell(row=rr, column=c, value=val)
    ws.cell(row=rr, column=4).font = Font(bold=True, color=GREEN_FONT)
band_rows(ws, hh + 1, rr, 4)
autosize(ws, {"A": 24, "B": 26, "C": 34, "D": 26})

# ═════════════════════════════════════════════════════════════════════════
# Sheet 2: Performance Comparison
# ═════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Performance Comparison")
write_title(ws2, "Performance Comparison", "100 Virtual Users · 60 Seconds · Same Docker Stack · Same Hardware · Same k6 Configuration")

hdr2 = 4
for c, h in enumerate(["Metric", "Before", "After", "Improvement"], start=1):
    ws2.cell(row=hdr2, column=c, value=h)
style_header_row(ws2, hdr2, 4)
perf_table = [
    ("Success Rate", "1.25%", "86.19% (100% excluding expected 409 duplicate-upload rejections)", "Significant Improvement"),
    ("Average Response Time", "1736.9 ms", "176.7 ms", "-89.8%"),
    ("P99 Latency", "32340.7 ms", "1628.8 ms", "-95.0%"),
    ("Login Success", "15.5% (29.1 s avg)", "100% (323.5 ms avg)", "Fully Resolved"),
    ("401 Errors", "1424", "0", "Eliminated"),
    ("500 Errors", "71", "0", "Eliminated"),
]
r = hdr2
for row in perf_table:
    r += 1
    for c, val in enumerate(row, start=1):
        ws2.cell(row=r, column=c, value=val).alignment = LEFT if c == 3 else CENTER
    ws2.cell(row=r, column=4).font = Font(bold=True, color=GREEN_FONT)
    ws2.row_dimensions[r].height = 22
band_rows(ws2, hdr2 + 1, r, 4)
autosize(ws2, {"A": 24, "B": 20, "C": 50, "D": 26})

# Chart data: Before vs After Response Time / Success Rate
chart_start = r + 3
ws2.cell(row=chart_start, column=1, value="Chart Data").font = SECTION_FONT
ch = chart_start + 1
for c, h in enumerate(["Metric", "Before", "After"], start=1):
    ws2.cell(row=ch, column=c, value=h)
style_header_row(ws2, ch, 3)
resp_metrics = [("Avg (ms)", cmp_metric("Avg Response Time (ms)")), ("P95 (ms)", cmp_metric("P95 Response Time (ms)")),
                ("P99 (ms)", cmp_metric("P99 Response Time (ms)"))]
succ_metric = cmp_metric("Success Rate (%)")
rr = ch
resp_first = rr + 1
for label, m in resp_metrics:
    rr += 1
    ws2.cell(row=rr, column=1, value=label)
    ws2.cell(row=rr, column=2, value=m["before"]).number_format = "#,##0.00"
    ws2.cell(row=rr, column=3, value=m["after"]).number_format = "#,##0.00"
resp_last = rr
rr += 1
succ_row = rr
ws2.cell(row=rr, column=1, value="Success %")
ws2.cell(row=rr, column=2, value=succ_metric["before"]).number_format = "#,##0.00"
ws2.cell(row=rr, column=3, value=succ_metric["after"]).number_format = "#,##0.00"
band_rows(ws2, ch + 1, rr, 3)

resp_chart = BarChart()
resp_chart.type = "col"
resp_chart.grouping = "clustered"
resp_chart.title = "Before vs After Response Time (ms)"
resp_chart.y_axis.title = "ms"
resp_chart.style = 10
resp_chart.add_data(Reference(ws2, min_col=2, max_col=3, min_row=ch, max_row=resp_last), titles_from_data=True)
resp_chart.set_categories(Reference(ws2, min_col=1, min_row=resp_first, max_row=resp_last))
resp_chart.width, resp_chart.height = 15, 9
ws2.add_chart(resp_chart, f"F{chart_start}")

succ_chart = BarChart()
succ_chart.type = "col"
succ_chart.grouping = "clustered"
succ_chart.title = "Before vs After Success Rate (%)"
succ_chart.y_axis.title = "%"
succ_chart.style = 12
succ_chart.add_data(Reference(ws2, min_col=2, max_col=3, min_row=ch, max_row=succ_row), titles_from_data=True)
succ_chart.set_categories(Reference(ws2, min_col=1, min_row=succ_row, max_row=succ_row))
succ_chart.width, succ_chart.height = 15, 9
ws2.add_chart(succ_chart, f"F{chart_start + 18}")

# ═════════════════════════════════════════════════════════════════════════
# Sheet 3: Before vs After Metrics
# ═════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Before vs After Metrics")
write_title(ws3, "Before vs After Metrics", "Full overall load-test metrics, pre- and post-fix")
hdr3 = 4
for c, h in enumerate(["Metric", "Before", "After", "% Change"], start=1):
    ws3.cell(row=hdr3, column=c, value=h)
style_header_row(ws3, hdr3, 4)
r = hdr3
for row in overall_cmp:
    r += 1
    ws3.cell(row=r, column=1, value=row["metric"])
    ws3.cell(row=r, column=2, value=row["before"]).number_format = "#,##0.00"
    ws3.cell(row=r, column=3, value=row["after"]).number_format = "#,##0.00"
    change = row["pct_change"]
    cc = ws3.cell(row=r, column=4)
    if change is not None:
        cc.value = change / 100
        cc.number_format = "+#,##0.00%;-#,##0.00%"
        good = change > 0 if row["direction"] == "higher_better" else change < 0
        cc.font = Font(bold=True, color=GREEN_FONT if good else RED_FONT)
    else:
        cc.value = "n/a"
band_rows(ws3, hdr3 + 1, r, 4)
autosize(ws3, {"A": 30, "B": 16, "C": 16, "D": 14})

# ═════════════════════════════════════════════════════════════════════════
# Sheet 4: Endpoint Performance
# ═════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Endpoint Performance")
write_title(ws4, "Endpoint Performance", "Per-endpoint latency and success rate, before vs after")
headers4 = ["Endpoint", "Method", "Before Avg (ms)", "After Avg (ms)", "Before P95 (ms)", "After P95 (ms)",
            "Before P99 (ms)", "After P99 (ms)", "Before Success %", "After Success %"]
hdr4 = 4
for c, h in enumerate(headers4, start=1):
    ws4.cell(row=hdr4, column=c, value=h)
style_header_row(ws4, hdr4, len(headers4))
r = hdr4
for e in endpoint_cmp:
    r += 1
    ws4.cell(row=r, column=1, value=e["endpoint"])
    ws4.cell(row=r, column=2, value=e["method"])
    ws4.cell(row=r, column=3, value=e["before_avg_ms"]).number_format = "#,##0.00"
    ws4.cell(row=r, column=4, value=e["after_avg_ms"]).number_format = "#,##0.00"
    ws4.cell(row=r, column=5, value=e["before_p95_ms"]).number_format = "#,##0.00"
    ws4.cell(row=r, column=6, value=e["after_p95_ms"]).number_format = "#,##0.00"
    ws4.cell(row=r, column=7, value=e["before_p99_ms"]).number_format = "#,##0.00"
    ws4.cell(row=r, column=8, value=e["after_p99_ms"]).number_format = "#,##0.00"
    ws4.cell(row=r, column=9, value=e["before_success_pct"]).number_format = "#,##0.00"
    ws4.cell(row=r, column=10, value=e["after_success_pct"]).number_format = "#,##0.00"
band_rows(ws4, hdr4 + 1, r, len(headers4))
if r > hdr4:
    pct_conditional_format(ws4, f"J{hdr4+1}:J{r}", 95)
autosize(ws4, {"A": 26})

# ═════════════════════════════════════════════════════════════════════════
# Sheet 5: HTTP Status Codes
# ═════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("HTTP Status Codes")
write_title(ws5, "HTTP Status Codes", "Response status code distribution, before vs after")
headers5 = ["Status Code", "Before Count", "Before %", "After Count", "After %"]
hdr5 = 4
for c, h in enumerate(headers5, start=1):
    ws5.cell(row=hdr5, column=c, value=h)
style_header_row(ws5, hdr5, 5)
STATUS_LABELS = {"0": "0 (Timeout/Network)", "200": "200 OK", "201": "201 Created", "401": "401 Unauthorized",
                  "409": "409 Conflict (dup. upload)", "429": "429 Rate Limited", "500": "500 Server Error"}
r = hdr5
error_rows = {}
for s in status_cmp:
    r += 1
    ws5.cell(row=r, column=1, value=STATUS_LABELS.get(s["status_code"], s["status_code"]))
    ws5.cell(row=r, column=2, value=s["before_count"]).number_format = "#,##0"
    ws5.cell(row=r, column=3, value=(s["before_pct"] / 100) if s["before_pct"] else 0).number_format = "0.00%"
    ws5.cell(row=r, column=4, value=s["after_count"]).number_format = "#,##0"
    ws5.cell(row=r, column=5, value=(s["after_pct"] / 100) if s["after_pct"] else 0).number_format = "0.00%"
    if s["status_code"] in ("401", "500"):
        error_rows[s["status_code"]] = r
band_rows(ws5, hdr5 + 1, r, 5)
autosize(ws5, {"A": 28})

before_pie = PieChart()
before_pie.title = "Before: Status Code Distribution"
before_pie.add_data(Reference(ws5, min_col=2, min_row=hdr5, max_row=r), titles_from_data=True)
before_pie.set_categories(Reference(ws5, min_col=1, min_row=hdr5 + 1, max_row=r))
before_pie.dataLabels = DataLabelList()
before_pie.dataLabels.showPercent = True
before_pie.width, before_pie.height = 13, 8
ws5.add_chart(before_pie, f"G{hdr5}")

after_pie = PieChart()
after_pie.title = "After: Status Code Distribution"
after_pie.add_data(Reference(ws5, min_col=4, min_row=hdr5, max_row=r), titles_from_data=True)
after_pie.set_categories(Reference(ws5, min_col=1, min_row=hdr5 + 1, max_row=r))
after_pie.dataLabels = DataLabelList()
after_pie.dataLabels.showPercent = True
after_pie.width, after_pie.height = 13, 8
ws5.add_chart(after_pie, f"G{hdr5 + 17}")

# HTTP Error Counts chart data (401 + 500 specifically)
err_chart_start = r + 3
ws5.cell(row=err_chart_start, column=1, value="HTTP Error Counts (401 + 500)").font = SECTION_FONT
ech = err_chart_start + 1
for c, h in enumerate(["Status", "Before", "After"], start=1):
    ws5.cell(row=ech, column=c, value=h)
style_header_row(ws5, ech, 3)
rr = ech
for code in ("401", "500"):
    rr += 1
    src = error_rows.get(code)
    before_val = ws5.cell(row=src, column=2).value if src else 0
    after_val = ws5.cell(row=src, column=4).value if src else 0
    ws5.cell(row=rr, column=1, value=f"{code} Errors")
    ws5.cell(row=rr, column=2, value=before_val).number_format = "#,##0"
    ws5.cell(row=rr, column=3, value=after_val).number_format = "#,##0"
band_rows(ws5, ech + 1, rr, 3)

err_chart = BarChart()
err_chart.type = "col"
err_chart.grouping = "clustered"
err_chart.title = "HTTP Error Counts: Before vs After"
err_chart.y_axis.title = "Count"
err_chart.style = 13
err_chart.add_data(Reference(ws5, min_col=2, max_col=3, min_row=ech, max_row=rr), titles_from_data=True)
err_chart.set_categories(Reference(ws5, min_col=1, min_row=ech + 1, max_row=rr))
err_chart.width, err_chart.height = 14, 8
ws5.add_chart(err_chart, f"G{err_chart_start}")

# ═════════════════════════════════════════════════════════════════════════
# Sheet 6: Resource Usage
# ═════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Resource Usage")
write_title(ws6, "Resource Usage", "docker stats averages/peaks, before vs after")
headers6 = ["Container", "Before Avg CPU %", "After Avg CPU %", "Before Max CPU %", "After Max CPU %",
            "Before Avg Mem (MB)", "After Avg Mem (MB)", "Before Max Mem (MB)", "After Max Mem (MB)"]
hdr6 = 4
for c, h in enumerate(headers6, start=1):
    ws6.cell(row=hdr6, column=c, value=h)
style_header_row(ws6, hdr6, len(headers6))
r = hdr6
for res in resource_cmp:
    r += 1
    ws6.cell(row=r, column=1, value=res["container"])
    ws6.cell(row=r, column=2, value=res["before_avg_cpu_pct"]).number_format = "#,##0.00"
    ws6.cell(row=r, column=3, value=res["after_avg_cpu_pct"]).number_format = "#,##0.00"
    ws6.cell(row=r, column=4, value=res["before_max_cpu_pct"]).number_format = "#,##0.00"
    ws6.cell(row=r, column=5, value=res["after_max_cpu_pct"]).number_format = "#,##0.00"
    ws6.cell(row=r, column=6, value=res["before_avg_mem_mb"]).number_format = "#,##0.00"
    ws6.cell(row=r, column=7, value=res["after_avg_mem_mb"]).number_format = "#,##0.00"
    ws6.cell(row=r, column=8, value=res["before_max_mem_mb"]).number_format = "#,##0.00"
    ws6.cell(row=r, column=9, value=res["after_max_mem_mb"]).number_format = "#,##0.00"
band_rows(ws6, hdr6 + 1, r, len(headers6))
autosize(ws6, {"A": 26})

cpu_chart = BarChart()
cpu_chart.type = "col"
cpu_chart.grouping = "clustered"
cpu_chart.title = "Avg CPU %: Before vs After (by container)"
cpu_chart.y_axis.title = "CPU %"
cpu_chart.style = 10
cpu_chart.add_data(Reference(ws6, min_col=2, max_col=3, min_row=hdr6, max_row=r), titles_from_data=True)
cpu_chart.set_categories(Reference(ws6, min_col=1, min_row=hdr6 + 1, max_row=r))
cpu_chart.width, cpu_chart.height = 16, 9
ws6.add_chart(cpu_chart, f"B{r + 3}")

# ═════════════════════════════════════════════════════════════════════════
# Sheet 7: Bottleneck Analysis
# ═════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Bottleneck Analysis")
write_title(ws7, "Bottleneck Analysis", "Primary and secondary bottlenecks identified during this fix")

ws7.cell(row=4, column=1, value="Primary Bottleneck").font = SECTION_FONT
ws7.merge_cells("A5:F5")
ws7["A5"] = ("bcrypt.checkpw() and bcrypt.hashpw() executed synchronously inside async authentication routes "
             "(api/core/security.py), blocking the single asyncio event loop for ~150-300ms per call. FastAPI runs "
             "a single uvicorn worker with one event loop, so every other in-flight request (DB queries, JWT, "
             "unrelated routes, other users' logins) queued behind whichever request was mid-bcrypt.")
ws7["A5"].alignment = LEFT
ws7.row_dimensions[5].height = 70

ws7.cell(row=7, column=1, value="Solution").font = SECTION_FONT
ws7.merge_cells("A8:F8")
ws7["A8"] = ("Moved password hashing and verification to worker threads using asyncio.to_thread(). Security "
             "unchanged: bcrypt algorithm unchanged, 12 rounds unchanged, hash format unchanged — existing "
             "password hashes verify identically.")
ws7["A8"].alignment = LEFT
ws7.row_dimensions[8].height = 50

ws7.cell(row=10, column=1, value="Secondary Bottlenecks (exposed once the primary bottleneck was fixed)").font = SECTION_FONT
hdr7 = 12
for c, h in enumerate(["Bottleneck", "Symptom", "Solution", "Change Type"], start=1):
    ws7.cell(row=hdr7, column=c, value=h)
style_header_row(ws7, hdr7, 4)
secondary = [
    ("PostgreSQL QueuePool exhaustion", "sqlalchemy.exc.TimeoutError: QueuePool limit ... reached, connection "
     "timed out — surfaced once bcrypt stopped serializing requests and 100 concurrent VUs reached the database "
     "for the first time", "Pool size increased 5/20 → 10/50 (DB_POOL_MIN_SIZE / DB_POOL_MAX_SIZE)",
     "Configuration-only"),
    ("Redis connection exhaustion", "redis.exceptions.ConnectionError: Too many connections — same mechanism, "
     "for the rate limiter's Redis client pool", "Pool size increased 10 → 50 (REDIS_POOL_MAX_SIZE)",
     "Configuration-only"),
    ("k6 test harness cookie persistence", "This k6 build's implicit per-VU cookie jar did not reliably persist "
     "the session cookie across iterations under concurrent load, producing a false 401 cascade unrelated to "
     "backend behavior (proven via direct curl and isolated single-VU/10-VU k6 probes)", "Script explicitly "
     "captures Set-Cookie from the login response and replays it as a Cookie header on every subsequent request",
     "Test-harness fix (not an application change)"),
]
r = hdr7
for row in secondary:
    r += 1
    for c, val in enumerate(row, start=1):
        ws7.cell(row=r, column=c, value=val).alignment = LEFT
    ws7.row_dimensions[r].height = 65
band_rows(ws7, hdr7 + 1, r, 4)
autosize(ws7, {"A": 26, "B": 45, "C": 40, "D": 22})

# ═════════════════════════════════════════════════════════════════════════
# Sheet 8: Optimizations Implemented
# ═════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Optimizations Implemented")
write_title(ws8, "Optimizations Implemented", "Concrete code changes made for this fix")
headers8 = ["File", "Change", "Why"]
hdr8 = 4
for c, h in enumerate(headers8, start=1):
    ws8.cell(row=hdr8, column=c, value=h)
style_header_row(ws8, hdr8, 3)
optimisations = [
    ("api/core/security.py", "hash_password() and verify_password() are now async, running the same "
     "bcrypt.hashpw/checkpw call (same 12 rounds, same $2b$ hash format) via asyncio.to_thread instead of "
     "directly on the event loop.", "Moves the CPU-bound bcrypt call off the event loop thread so it no longer "
     "blocks every other in-flight request. Zero change to the hash algorithm, cost factor, or resulting hash."),
    ("api/auth/service.py", "5 call sites (register_user, login_user, reset_password, change_password x2) "
     "updated to await hash_password(...) / await verify_password(...).", "Required by hash_password/"
     "verify_password becoming async; no logic changed."),
    ("api/core/config.py + api/.env.example", "DB_POOL_MIN_SIZE 5→10, DB_POOL_MAX_SIZE 20→50.",
     "The 20-connection pool was adequate only while bcrypt serialized requests; once removed, 100 concurrent "
     "VUs exhausted it. 50 stays well within Postgres's default max_connections=100."),
    ("api/core/config.py", "REDIS_POOL_MAX_SIZE 10→50.", "Same reasoning as the DB pool, for the rate "
     "limiter's Redis client connections."),
    ("performance/k6/baseline_load_test.js", "Login now captures Set-Cookie values from the response and "
     "attaches them explicitly as a Cookie header on every subsequent request for that VU.", "Test-harness fix: "
     "the implicit per-VU cookie jar did not reliably persist across iterations under load in this k6 build. "
     "Traffic mix, VU/stage profile, sleep, and thresholds are all unchanged."),
]
r = hdr8
for row in optimisations:
    r += 1
    for c, val in enumerate(row, start=1):
        ws8.cell(row=r, column=c, value=val).alignment = LEFT
    ws8.row_dimensions[r].height = 85
band_rows(ws8, hdr8 + 1, r, 3)
autosize(ws8, {"A": 30, "B": 55, "C": 55})

not_changed_row = r + 3
ws8.cell(row=not_changed_row, column=1, value="Explicitly NOT Changed").font = SECTION_FONT
ws8.merge_cells(f"A{not_changed_row + 1}:C{not_changed_row + 3}")
ws8.cell(row=not_changed_row + 1, column=1, value=(
    "Password hashing algorithm/cost factor (still bcrypt, 12 rounds) — security unchanged. No uvicorn "
    "--workers/process-model change — the single-event-loop architecture was preserved. No caching layer "
    "added. No new services/queues introduced. api/inference/** (AI pipeline) untouched — it was already "
    "correctly using asyncio.to_thread throughout."
)).alignment = LEFT

# ═════════════════════════════════════════════════════════════════════════
# Sheet 9: Regression Testing
# ═════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Regression Testing")
write_title(ws9, "Regression Testing", "Verification that no functionality, behavior, or security regressed")
hdr9 = 4
for c, h in enumerate(["Check", "Result", "Detail"], start=1):
    ws9.cell(row=hdr9, column=c, value=h)
style_header_row(ws9, hdr9, 3)
regression_rows = [
    ("Backend Tests", "191 / 191 Passed", "Full api/tests/ suite, including auth, register, login, "
     "change-password, reset-password, and every other slice's tests"),
    ("TypeScript", "Passed", "tsc -b — clean, no type errors"),
    ("ESLint", "Passed", "eslint . — 0 errors (pre-existing warnings only, unrelated to this change)"),
    ("Vite Build", "Passed", "npm run build — production build succeeds"),
    ("Authentication Behavior", "Unchanged", "Same hash format, same constant-time-shaped dummy-hash comparison "
     "on login, same session-revoke-on-password-change behavior, same rate-limit codes/thresholds"),
    ("Security", "Unchanged", "bcrypt cost factor unchanged (12 rounds); no plaintext password ever touches "
     "disk or logs; asyncio.to_thread introduces no new I/O or network exposure"),
]
r = hdr9
for label, result, detail in regression_rows:
    r += 1
    ws9.cell(row=r, column=1, value=label).font = LABEL_FONT
    rc = ws9.cell(row=r, column=2, value=result)
    rc.font = Font(bold=True, color=GREEN_FONT)
    rc.fill = GREEN_FILL
    rc.alignment = CENTER
    ws9.cell(row=r, column=3, value=detail).alignment = LEFT
    ws9.row_dimensions[r].height = 40
band_rows(ws9, hdr9 + 1, r, 3)
autosize(ws9, {"A": 26, "B": 20, "C": 60})

# ═════════════════════════════════════════════════════════════════════════
# Sheet 10: Charts
# ═════════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("Charts")
write_title(ws10, "Charts", "Consolidated before vs after comparison charts")

# Chart data table (single source for every chart on this sheet)
hdr10 = 4
for c, h in enumerate(["Metric", "Before", "After"], start=1):
    ws10.cell(row=hdr10, column=c, value=h)
style_header_row(ws10, hdr10, 3)
chart10_metrics = [
    ("Success Rate (%)", cmp_metric("Success Rate (%)")["before"], cmp_metric("Success Rate (%)")["after"]),
    ("Avg Response Time (ms)", cmp_metric("Avg Response Time (ms)")["before"], cmp_metric("Avg Response Time (ms)")["after"]),
    ("P99 Latency (ms)", cmp_metric("P99 Response Time (ms)")["before"], cmp_metric("P99 Response Time (ms)")["after"]),
    ("Login Success (%)", 15.48, 100.0),
    ("401 Errors", 1424, 0),
    ("500 Errors", 71, 0),
]
row_index = {}
r = hdr10
for label, before, after in chart10_metrics:
    r += 1
    ws10.cell(row=r, column=1, value=label)
    ws10.cell(row=r, column=2, value=before).number_format = "#,##0.00"
    ws10.cell(row=r, column=3, value=after).number_format = "#,##0.00"
    row_index[label] = r
band_rows(ws10, hdr10 + 1, r, 3)
autosize(ws10, {"A": 24, "B": 14, "C": 14})


def single_metric_chart(label: str, title: str, y_title: str, anchor: str, style: int):
    row = row_index[label]
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = y_title
    chart.style = style
    chart.add_data(Reference(ws10, min_col=2, max_col=3, min_row=hdr10, max_row=row), titles_from_data=True)
    chart.set_categories(Reference(ws10, min_col=1, min_row=row, max_row=row))
    chart.width, chart.height = 13, 8
    ws10.add_chart(chart, anchor)


single_metric_chart("Success Rate (%)", "Success Rate: Before vs After", "%", "E4", 10)
single_metric_chart("Avg Response Time (ms)", "Average Response Time: Before vs After", "ms", "E20", 11)
single_metric_chart("P99 Latency (ms)", "P99 Latency: Before vs After", "ms", "E36", 12)
single_metric_chart("Login Success (%)", "Login Success: Before vs After", "%", "E52", 13)

err_row_first, err_row_last = row_index["401 Errors"], row_index["500 Errors"]
error_counts_chart = BarChart()
error_counts_chart.type = "col"
error_counts_chart.grouping = "clustered"
error_counts_chart.title = "HTTP Error Counts: Before vs After"
error_counts_chart.y_axis.title = "Count"
error_counts_chart.style = 14
error_counts_chart.add_data(Reference(ws10, min_col=2, max_col=3, min_row=hdr10, max_row=err_row_last), titles_from_data=True)
error_counts_chart.set_categories(Reference(ws10, min_col=1, min_row=err_row_first, max_row=err_row_last))
error_counts_chart.width, error_counts_chart.height = 13, 8
ws10.add_chart(error_counts_chart, "L4")

# Before vs After Response Time (avg/p95/p99 grouped) and Before vs After Success Rate,
# reusing the same underlying data already on "Performance Comparison" for consistency.
resp_group_chart = BarChart()
resp_group_chart.type = "col"
resp_group_chart.grouping = "clustered"
resp_group_chart.title = "Before vs After Response Time (Avg / P99)"
resp_group_chart.y_axis.title = "ms"
resp_group_chart.style = 11
avg_row, p99_row = row_index["Avg Response Time (ms)"], row_index["P99 Latency (ms)"]
resp_group_chart.add_data(Reference(ws10, min_col=2, max_col=3, min_row=hdr10, max_row=p99_row), titles_from_data=True)
resp_group_chart.set_categories(Reference(ws10, min_col=1, min_row=avg_row, max_row=p99_row))
resp_group_chart.width, resp_group_chart.height = 13, 8
ws10.add_chart(resp_group_chart, "L20")

# ═════════════════════════════════════════════════════════════════════════
# Sheet 11: Final Assessment
# ═════════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("Final Assessment")
write_title(ws11, "Final Assessment", "Concluding summary")

verdict_cell(ws11, "B4", True, "PASS", size=16)
ws11["A4"] = "Overall Result"
ws11["A4"].font = LABEL_FONT
ws11.row_dimensions[4].height = 24

conclusion_points = [
    "The primary event-loop bottleneck (bcrypt executing synchronously inside async authentication routes) "
    "was eliminated by moving password hashing/verification onto worker threads via asyncio.to_thread().",
    "Database and Redis connection pool limits were tuned (Postgres 5/20 → 10/50, Redis 10 → 50) after "
    "being exposed as secondary bottlenecks once the primary block was removed and real concurrency reached "
    "them for the first time.",
    "Authentication performance improved dramatically without reducing security: bcrypt algorithm, cost "
    "factor (12 rounds), and hash format are all unchanged; existing password hashes verify identically.",
    "No functional regressions were introduced: API contracts, response shapes, and business rules are "
    "unchanged.",
    "All backend tests continue to pass (191/191), alongside clean TypeScript, ESLint, and Vite build results.",
    "The optimized application is significantly more capable of sustaining 100 concurrent virtual users: "
    "success rate rose from 1.25% to 86.19% (100% excluding expected 409 duplicate-upload rejections), average "
    "response time fell 89.8% (1,736.9ms → 176.7ms), P99 latency fell 95.0% (32,340.7ms → 1,628.8ms), "
    "and both 401 and 500 errors were eliminated entirely (1,424 → 0 and 71 → 0, respectively).",
]
r = 6
ws11.cell(row=r, column=1, value="Conclusion").font = SECTION_FONT
r += 1
for point in conclusion_points:
    ws11.merge_cells(f"A{r}:F{r}")
    cell = ws11.cell(row=r, column=1, value=f"• {point}")
    cell.alignment = LEFT
    ws11.row_dimensions[r].height = 46
    r += 1

ws11.column_dimensions["A"].width = 20
for col in "BCDEF":
    ws11.column_dimensions[col].width = 16

# ── Reorder & apply baseline-matching sheet styling (gridlines off) ───────
order = [
    "Executive Summary", "Performance Comparison", "Before vs After Metrics", "Endpoint Performance",
    "HTTP Status Codes", "Resource Usage", "Bottleneck Analysis", "Optimizations Implemented",
    "Regression Testing", "Charts", "Final Assessment",
]
wb._sheets = [wb[name] for name in order]
for ws_ in wb.worksheets:
    ws_.sheet_view.showGridLines = False

wb.save(OUT_XLSX)
print("Wrote:", OUT_XLSX)
print("Original baseline workbook untouched at:", ORIGINAL_XLSX)
