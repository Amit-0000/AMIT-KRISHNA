"""Builds performance/VoiceGuard_Baseline_Load_Test_Report.xlsx from the
parsed baseline test results (baseline_results.json / .csv) plus the
resource-usage sampler CSV. Pure openpyxl, no pandas/numpy.
"""
from __future__ import annotations

import csv
import json
import math
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).parent
RESULTS_JSON = ROOT / "baseline_results.json"
RESULTS_CSV = ROOT / "baseline_results.csv"
RESOURCE_CSV = ROOT / "results" / "resource_usage.csv"
OUT_XLSX = ROOT / "VoiceGuard_Baseline_Load_Test_Report.xlsx"

# ─── Palette / styles ──────────────────────────────────────────────────────
NAVY = "1F2937"
ACCENT = "2563EB"
LIGHT_BAND = "F3F4F6"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
GREEN_FONT = "1E7B34"
YELLOW = "FFEB9C"
YELLOW_FONT = "9C6500"
RED = "FFC7CE"
RED_FONT = "9C0006"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=16)
SUBTITLE_FONT = Font(italic=True, color="6B7280", size=10)
SECTION_FONT = Font(bold=True, color=ACCENT, size=12)
LABEL_FONT = Font(bold=True, color=NAVY)
BAND_FILL = PatternFill("solid", fgColor=LIGHT_BAND)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws: Worksheet, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def band_rows(ws: Worksheet, start_row: int, end_row: int, n_cols: int):
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = BAND_FILL
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = BORDER


def autosize(ws: Worksheet, widths: dict[str, int] | None = None, min_width=10, max_width=60):
    widths = widths or {}
    col_max = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            length = len(str(cell.value))
            col_max[col] = max(col_max.get(col, 0), length)
    for col, length in col_max.items():
        w = widths.get(col, max(min_width, min(max_width, length + 3)))
        ws.column_dimensions[col].width = w


def write_title(ws: Worksheet, title: str, subtitle: str = ""):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT


def response_time_conditional_format(ws: Worksheet, cell_range: str):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["200"], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FONT)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between", formula=["200", "500"], fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FONT)
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["500"], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)),
    )


def error_rate_conditional_format(ws: Worksheet, cell_range: str):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["1"], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)),
    )


# ─── Load data ──────────────────────────────────────────────────────────────
results = json.loads(RESULTS_JSON.read_text(encoding="utf-8"))
overall = results["overall_performance"]
endpoints = results["endpoint_performance"]
status_codes = results["http_status_codes"]
distribution = results["response_time_distribution"]
resource_summary = results["resource_usage_summary"]
meta = results["test_metadata"]

with RESULTS_CSV.open(encoding="utf-8") as f:
    raw_requests = list(csv.DictReader(f))

resource_rows = []
if RESOURCE_CSV.exists():
    with RESOURCE_CSV.open(encoding="utf-8") as f:
        resource_rows = list(csv.DictReader(f))

CONTAINER_LABELS = {
    "voiceguard-backend-1": "Backend (FastAPI)",
    "voiceguard-postgres-1": "PostgreSQL",
    "voiceguard-redis-1": "Redis",
    "voiceguard-frontend-1": "Frontend (Vite/React)",
}

TEST_START = meta["test_window_start"]
TEST_END = meta["test_window_end"]


def parse_ts(s: str) -> datetime:
    return datetime.fromisoformat(s.replace("Z", "+00:00"))


# ═════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Executive Summary ─────────────────────────────────────────────
ws = wb.active
ws.title = "Executive Summary"
write_title(ws, "VoiceGuard — Baseline Load Test Report", "Executive Summary")

result_flag = "FAIL" if overall["error_rate_pct"] > 1 else "PASS"
flag_fill = RED if result_flag == "FAIL" else GREEN
flag_font_color = RED_FONT if result_flag == "FAIL" else GREEN_FONT

ws["A4"] = "Overall Result"
ws["A4"].font = LABEL_FONT
ws["B4"] = result_flag
ws["B4"].font = Font(bold=True, size=14, color=flag_font_color)
ws["B4"].fill = PatternFill("solid", fgColor=flag_fill)
ws["B4"].alignment = CENTER

rows = [
    ("Project Name", "VoiceGuard"),
    ("Test Type", "Baseline Load Test"),
    ("Date & Time", parse_ts(TEST_START).strftime("%Y-%m-%d %H:%M:%S UTC") if TEST_START else "N/A"),
    ("Duration", f"1 Minute (measured: {overall['duration_s']}s)"),
    ("Concurrent Users", overall["virtual_users"]),
    ("Environment — OS", "Windows 11 Home Single Language (10.0.26200)"),
    ("Environment — CPU", "12th Gen Intel Core i7-12700H (14 cores / 20 threads)"),
    ("Environment — RAM", "31.7 GB"),
    ("Environment — Python Version", "3.12.13 (backend container, python:3.12-slim)"),
    ("Backend", "FastAPI (uvicorn, single worker) — api.main:app"),
    ("Database", "PostgreSQL 16 (postgres:16-alpine, Docker)"),
    ("Cache", "Redis 7 (redis:7-alpine, Docker)"),
    ("Frontend", "React (Vite dev server, node:20-alpine)"),
    ("Model Version", "Configured: lcnn v1 (checkpoints/best.pt) — NOT loaded: file missing and torch not installed in this backend image; /predict returns 503"),
    ("Checkpoint SHA256", "deepfake_cnn.pth (present, unregistered AudioCNN checkpoint): "
        "0fbe937e222a057ea56457334b41558f56a5c1472fa720b7083aa1c3bc70d4e7"),
]
r = 6
for label, value in rows:
    ws.cell(row=r, column=1, value=label).font = LABEL_FONT
    ws.cell(row=r, column=2, value=value).alignment = LEFT
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2).border = BORDER
    if (r % 2) == 1:
        ws.cell(row=r, column=1).fill = BAND_FILL
        ws.cell(row=r, column=2).fill = BAND_FILL
    r += 1

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 95
for rr in range(6, r):
    ws.row_dimensions[rr].height = 18

# ── Sheet 2: Overall Performance ───────────────────────────────────────────
ws2 = wb.create_sheet("Overall Performance")
write_title(ws2, "Overall Performance", f"Window: {TEST_START} → {TEST_END}")

metric_rows = [
    ("Virtual Users", overall["virtual_users"], "#,##0"),
    ("Duration (s)", overall["duration_s"], "#,##0.0"),
    ("Total Requests", overall["total_requests"], "#,##0"),
    ("Successful Requests", overall["successful_requests"], "#,##0"),
    ("Failed Requests", overall["failed_requests"], "#,##0"),
    ("Requests Per Second (RPS)", overall["requests_per_second"], "#,##0.00"),
    ("Average Response Time (ms)", overall["avg_response_time_ms"], "#,##0.00"),
    ("Median Response Time (ms)", overall["median_response_time_ms"], "#,##0.00"),
    ("Minimum Response Time (ms)", overall["min_response_time_ms"], "#,##0.00"),
    ("Maximum Response Time (ms)", overall["max_response_time_ms"], "#,##0.00"),
    ("P90 Response Time (ms)", overall["p90_response_time_ms"], "#,##0.00"),
    ("P95 Response Time (ms)", overall["p95_response_time_ms"], "#,##0.00"),
    ("P99 Response Time (ms)", overall["p99_response_time_ms"], "#,##0.00"),
    ("Error Rate (%)", overall["error_rate_pct"], "#,##0.00"),
    ("Success Rate (%)", overall["success_rate_pct"], "#,##0.00"),
    ("Data Sent (KB)", round(overall["data_sent_bytes"] / 1024, 1) if overall["data_sent_bytes"] else 0, "#,##0.0"),
    ("Data Received (KB)", round(overall["data_received_bytes"] / 1024, 1) if overall["data_received_bytes"] else 0, "#,##0.0"),
]
ws2.cell(row=4, column=1, value="Metric")
ws2.cell(row=4, column=2, value="Value")
style_header_row(ws2, 4, 2)
r = 5
avg_rt_row = None
error_rate_row = None
for label, value, fmt in metric_rows:
    ws2.cell(row=r, column=1, value=label)
    vc = ws2.cell(row=r, column=2, value=value)
    vc.number_format = fmt
    if label == "Average Response Time (ms)":
        avg_rt_row = r
    if label == "Error Rate (%)":
        error_rate_row = r
    r += 1
band_rows(ws2, 5, r - 1, 2)
autosize(ws2, {"A": 32, "B": 16})
if avg_rt_row:
    response_time_conditional_format(ws2, f"B{avg_rt_row}")
if error_rate_row:
    error_rate_conditional_format(ws2, f"B{error_rate_row}")

# Percentile mini-table + chart
perc_start = r + 2
ws2.cell(row=perc_start, column=1, value="Response Time Percentiles").font = SECTION_FONT
perc_header = perc_start + 1
ws2.cell(row=perc_header, column=1, value="Percentile")
ws2.cell(row=perc_header, column=2, value="ms")
style_header_row(ws2, perc_header, 2)
perc_data = [
    ("Min", overall["min_response_time_ms"]),
    ("Median", overall["median_response_time_ms"]),
    ("P90", overall["p90_response_time_ms"]),
    ("P95", overall["p95_response_time_ms"]),
    ("P99", overall["p99_response_time_ms"]),
    ("Max", overall["max_response_time_ms"]),
]
for i, (label, val) in enumerate(perc_data):
    rr = perc_header + 1 + i
    ws2.cell(row=rr, column=1, value=label)
    ws2.cell(row=rr, column=2, value=val).number_format = "#,##0.00"
band_rows(ws2, perc_header + 1, perc_header + len(perc_data), 2)

perc_chart = BarChart()
perc_chart.title = "Response Time Percentiles (ms)"
perc_chart.y_axis.title = "ms"
perc_chart.style = 10
data_ref = Reference(ws2, min_col=2, min_row=perc_header, max_row=perc_header + len(perc_data))
cats_ref = Reference(ws2, min_col=1, min_row=perc_header + 1, max_row=perc_header + len(perc_data))
perc_chart.add_data(data_ref, titles_from_data=True)
perc_chart.set_categories(cats_ref)
perc_chart.width, perc_chart.height = 14, 8
ws2.add_chart(perc_chart, f"D{perc_start}")

# Time-bucketed RPS + Avg Response Time series (5s buckets)
ts_col_start = perc_header + len(perc_data) + 3
ws2.cell(row=ts_col_start, column=1, value="Requests Over Time (5s buckets)").font = SECTION_FONT
ts_header = ts_col_start + 1
for c, label in enumerate(["Bucket Start", "Requests", "RPS", "Avg Response (ms)"], start=1):
    ws2.cell(row=ts_header, column=c, value=label)
style_header_row(ws2, ts_header, 4)

if raw_requests:
    times = [parse_ts(rr["timestamp"]) for rr in raw_requests]
    t_min = min(times)
    bucket_s = 5
    buckets: dict[int, list[float]] = defaultdict(list)
    for rr, t in zip(raw_requests, times):
        idx = int((t - t_min).total_seconds() // bucket_s)
        buckets[idx].append(float(rr["response_time_ms"]))
    max_idx = max(buckets)
    rr_row = ts_header
    for idx in range(max_idx + 1):
        vals = buckets.get(idx, [])
        rr_row += 1
        ws2.cell(row=rr_row, column=1, value=f"{idx*bucket_s}-{(idx+1)*bucket_s}s")
        ws2.cell(row=rr_row, column=2, value=len(vals))
        ws2.cell(row=rr_row, column=3, value=round(len(vals) / bucket_s, 2))
        ws2.cell(row=rr_row, column=4, value=round(statistics.mean(vals), 2) if vals else 0)
    band_rows(ws2, ts_header + 1, rr_row, 4)

    rps_chart = LineChart()
    rps_chart.title = "Requests Per Second Over Time"
    rps_chart.y_axis.title = "RPS"
    rps_chart.style = 12
    data_ref = Reference(ws2, min_col=3, min_row=ts_header, max_row=rr_row)
    cats_ref = Reference(ws2, min_col=1, min_row=ts_header + 1, max_row=rr_row)
    rps_chart.add_data(data_ref, titles_from_data=True)
    rps_chart.set_categories(cats_ref)
    rps_chart.width, rps_chart.height = 16, 8
    ws2.add_chart(rps_chart, f"D{perc_start + 18}")

    art_chart = LineChart()
    art_chart.title = "Average Response Time Over Time"
    art_chart.y_axis.title = "ms"
    art_chart.style = 13
    data_ref = Reference(ws2, min_col=4, min_row=ts_header, max_row=rr_row)
    art_chart.add_data(data_ref, titles_from_data=True)
    art_chart.set_categories(cats_ref)
    art_chart.width, art_chart.height = 16, 8
    ws2.add_chart(art_chart, f"D{perc_start + 34}")

autosize(ws2, {"A": 32, "B": 16, "C": 12, "D": 20})

# ── Sheet 3: Endpoint Performance ──────────────────────────────────────────
ws3 = wb.create_sheet("Endpoint Performance")
write_title(ws3, "Endpoint Performance", "Per-endpoint latency and reliability during the baseline run")
headers3 = ["Endpoint", "HTTP Method", "Requests", "Average Response (ms)", "Minimum", "Maximum", "P95", "P99", "Success %", "Failure %"]
hdr_row3 = 4
for c, h in enumerate(headers3, start=1):
    ws3.cell(row=hdr_row3, column=c, value=h)
style_header_row(ws3, hdr_row3, len(headers3))
r = hdr_row3
for e in endpoints:
    r += 1
    ws3.cell(row=r, column=1, value=e["endpoint"])
    ws3.cell(row=r, column=2, value=e["method"])
    ws3.cell(row=r, column=3, value=e["requests"]).number_format = "#,##0"
    ws3.cell(row=r, column=4, value=e["avg_response_ms"]).number_format = "#,##0.00"
    ws3.cell(row=r, column=5, value=e["min_ms"]).number_format = "#,##0.00"
    ws3.cell(row=r, column=6, value=e["max_ms"]).number_format = "#,##0.00"
    ws3.cell(row=r, column=7, value=e["p95_ms"]).number_format = "#,##0.00"
    ws3.cell(row=r, column=8, value=e["p99_ms"]).number_format = "#,##0.00"
    ws3.cell(row=r, column=9, value=e["success_pct"]).number_format = "#,##0.00"
    ws3.cell(row=r, column=10, value=e["failure_pct"]).number_format = "#,##0.00"
band_rows(ws3, hdr_row3 + 1, r, len(headers3))
if r > hdr_row3:
    response_time_conditional_format(ws3, f"D{hdr_row3+1}:D{r}")
    error_rate_conditional_format(ws3, f"J{hdr_row3+1}:J{r}")
autosize(ws3, {"A": 28})

# ── Sheet 4: HTTP Status Codes ─────────────────────────────────────────────
ws4 = wb.create_sheet("HTTP Status Codes")
write_title(ws4, "HTTP Status Codes", "Distribution of response status codes across all requests")
hdr_row4 = 4
for c, h in enumerate(["Status Code", "Count", "Percentage"], start=1):
    ws4.cell(row=hdr_row4, column=c, value=h)
style_header_row(ws4, hdr_row4, 3)
STATUS_LABELS = {"0": "0 (Network Error / Timeout)", "200": "200 OK", "201": "201 Created", "400": "400 Bad Request",
                  "401": "401 Unauthorized", "403": "403 Forbidden", "404": "404 Not Found", "429": "429 Too Many Requests",
                  "500": "500 Internal Server Error"}
r = hdr_row4
for s in status_codes:
    r += 1
    ws4.cell(row=r, column=1, value=STATUS_LABELS.get(s["status_code"], s["status_code"]))
    ws4.cell(row=r, column=2, value=s["count"]).number_format = "#,##0"
    ws4.cell(row=r, column=3, value=s["percentage"] / 100).number_format = "0.00%"
band_rows(ws4, hdr_row4 + 1, r, 3)
autosize(ws4, {"A": 30})

status_chart = PieChart()
status_chart.title = "HTTP Status Code Distribution"
data_ref = Reference(ws4, min_col=2, min_row=hdr_row4, max_row=r)
cats_ref = Reference(ws4, min_col=1, min_row=hdr_row4 + 1, max_row=r)
status_chart.add_data(data_ref, titles_from_data=True)
status_chart.set_categories(cats_ref)
status_chart.dataLabels = DataLabelList()
status_chart.dataLabels.showPercent = True
status_chart.width, status_chart.height = 16, 10
ws4.add_chart(status_chart, f"E{hdr_row4}")

# ── Sheet 5: Resource Usage ────────────────────────────────────────────────
ws5 = wb.create_sheet("Resource Usage")
write_title(ws5, "Resource Usage", "docker stats sampled every ~5s across the test window")
headers5 = ["Timestamp", "Backend CPU %", "Backend Memory (MB)", "PostgreSQL CPU %", "PostgreSQL Memory (MB)",
            "Redis CPU %", "Redis Memory (MB)", "Frontend Memory (MB)"]
hdr_row5 = 4
for c, h in enumerate(headers5, start=1):
    ws5.cell(row=hdr_row5, column=c, value=h)
style_header_row(ws5, hdr_row5, len(headers5))

by_ts: dict[str, dict] = defaultdict(dict)
for row in resource_rows:
    by_ts[row["timestamp"]][row["container"]] = row

r = hdr_row5
for ts in sorted(by_ts.keys()):
    row = by_ts[ts]
    r += 1
    ws5.cell(row=r, column=1, value=ts)
    b = row.get("voiceguard-backend-1", {})
    p = row.get("voiceguard-postgres-1", {})
    rd = row.get("voiceguard-redis-1", {})
    fe = row.get("voiceguard-frontend-1", {})
    ws5.cell(row=r, column=2, value=float(b.get("cpu_pct", 0))).number_format = "#,##0.00"
    ws5.cell(row=r, column=3, value=float(b.get("mem_usage_mb", 0))).number_format = "#,##0.00"
    ws5.cell(row=r, column=4, value=float(p.get("cpu_pct", 0))).number_format = "#,##0.00"
    ws5.cell(row=r, column=5, value=float(p.get("mem_usage_mb", 0))).number_format = "#,##0.00"
    ws5.cell(row=r, column=6, value=float(rd.get("cpu_pct", 0))).number_format = "#,##0.00"
    ws5.cell(row=r, column=7, value=float(rd.get("mem_usage_mb", 0))).number_format = "#,##0.00"
    ws5.cell(row=r, column=8, value=float(fe.get("mem_usage_mb", 0))).number_format = "#,##0.00"
band_rows(ws5, hdr_row5 + 1, r, len(headers5))
autosize(ws5, {"A": 22})

if r > hdr_row5:
    cpu_chart = LineChart()
    cpu_chart.title = "Backend CPU Usage (%)"
    cpu_chart.y_axis.title = "CPU %"
    cpu_chart.style = 12
    data_ref = Reference(ws5, min_col=2, min_row=hdr_row5, max_row=r)
    cats_ref = Reference(ws5, min_col=1, min_row=hdr_row5 + 1, max_row=r)
    cpu_chart.add_data(data_ref, titles_from_data=True)
    cpu_chart.set_categories(cats_ref)
    cpu_chart.width, cpu_chart.height = 18, 9
    ws5.add_chart(cpu_chart, f"J{hdr_row5}")

    mem_chart = LineChart()
    mem_chart.title = "Backend Memory Usage (MB)"
    mem_chart.y_axis.title = "MB"
    mem_chart.style = 13
    data_ref = Reference(ws5, min_col=3, min_row=hdr_row5, max_row=r)
    mem_chart.add_data(data_ref, titles_from_data=True)
    mem_chart.set_categories(cats_ref)
    mem_chart.width, mem_chart.height = 18, 9
    ws5.add_chart(mem_chart, f"J{hdr_row5 + 20}")

# ── Sheet 6: Response Time Distribution ────────────────────────────────────
ws6 = wb.create_sheet("Response Time Distribution")
write_title(ws6, "Response Time Distribution", "All requests bucketed by response time")
hdr_row6 = 4
for c, h in enumerate(["Time Range (ms)", "Request Count"], start=1):
    ws6.cell(row=hdr_row6, column=c, value=h)
style_header_row(ws6, hdr_row6, 2)
r = hdr_row6
for d in distribution:
    r += 1
    ws6.cell(row=r, column=1, value=d["time_range_ms"])
    ws6.cell(row=r, column=2, value=d["request_count"]).number_format = "#,##0"
band_rows(ws6, hdr_row6 + 1, r, 2)
autosize(ws6, {"A": 20})

dist_chart = BarChart()
dist_chart.title = "Response Time Distribution"
dist_chart.y_axis.title = "Request Count"
dist_chart.x_axis.title = "Response Time (ms)"
dist_chart.style = 11
data_ref = Reference(ws6, min_col=2, min_row=hdr_row6, max_row=r)
cats_ref = Reference(ws6, min_col=1, min_row=hdr_row6 + 1, max_row=r)
dist_chart.add_data(data_ref, titles_from_data=True)
dist_chart.set_categories(cats_ref)
dist_chart.width, dist_chart.height = 16, 10
ws6.add_chart(dist_chart, "D4")

# ── Sheet 7: Raw Request Log ───────────────────────────────────────────────
ws7 = wb.create_sheet("Raw Request Log")
write_title(ws7, "Raw Request Log", f"{len(raw_requests):,} requests captured from k6 JSON output")
headers7 = ["Timestamp", "Virtual User", "Endpoint", "Method", "Status Code", "Response Time (ms)", "Success"]
hdr_row7 = 4
for c, h in enumerate(headers7, start=1):
    ws7.cell(row=hdr_row7, column=c, value=h)
style_header_row(ws7, hdr_row7, len(headers7))
r = hdr_row7
for rr in raw_requests:
    r += 1
    ws7.cell(row=r, column=1, value=rr["timestamp"])
    ws7.cell(row=r, column=2, value=int(rr["virtual_user"]) if rr["virtual_user"] else None)
    ws7.cell(row=r, column=3, value=rr["endpoint"])
    ws7.cell(row=r, column=4, value=rr["method"])
    ws7.cell(row=r, column=5, value=rr["status_code"])
    ws7.cell(row=r, column=6, value=float(rr["response_time_ms"])).number_format = "#,##0.000"
    ws7.cell(row=r, column=7, value="Yes" if rr["success"] == "True" else "No")
band_rows(ws7, hdr_row7 + 1, r, len(headers7))
if r > hdr_row7:
    ws7.conditional_formatting.add(
        f"G{hdr_row7+1}:G{r}",
        CellIsRule(operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)),
    )
autosize(ws7, {"A": 26, "C": 26})
ws7.freeze_panes = f"A{hdr_row7+1}"

# ── Sheet 8: Test Configuration ────────────────────────────────────────────
ws8 = wb.create_sheet("Test Configuration")
write_title(ws8, "Test Configuration", "k6 scenario definition used to produce this report")
config_rows = [
    ("Scenario Name", "baseline (ramping-vus executor)"),
    ("Virtual Users", "100 (peak)"),
    ("Duration", "60s total (10s ramp-up + 40s hold + 10s ramp-down)"),
    ("Ramp-up", "0 → 100 VUs over 10s"),
    ("Ramp-down", "100 → 0 VUs over 10s (5s graceful stop)"),
    ("Thresholds", "http_req_duration p(95) < 1000ms; http_req_failed rate < 5% (both breached — see Performance Assessment)"),
    ("Think Time", "0.5s – 1.5s uniform random sleep between iterations per VU"),
    ("Tool Version (k6)", "k6 v2.1.0 (grafana/k6 Docker image, commit 83a87a41e2, go1.26.4, linux/amd64)"),
    ("Traffic Mix", "40% GET /api/v1/scans, 25% GET /api/v1/user/profile, 20% GET /api/v1/scans/{id}, 15% POST /api/v1/scans (capped 5/VU)"),
    ("Test Users", "120 pre-seeded, pre-verified accounts (loadtest_user000..119@example.com), bypassing the register/email-verify flow"),
    ("Rate-limit note", "RATE_LIMIT_LOGIN_PER_HOUR_PER_IP raised from 10 to 5000 for this run only, via docker-compose.loadtest.yml override — "
        "all k6 traffic originates from a single container IP, so the default per-IP limit (correct for real deployments with distinct client IPs) "
        "would otherwise block ~90% of virtual users before the app itself was ever load tested. Reverted after the run."),
    ("Execution Environment", "docker compose (postgres:16-alpine, redis:7-alpine, backend built from api/Dockerfile, node:20-alpine frontend); "
        "k6 run from the official grafana/k6 Docker image on the voiceguard_default bridge network, hitting the backend container directly (http://backend:8000)"),
]
r = 4
for label, value in config_rows:
    ws8.cell(row=r, column=1, value=label).font = LABEL_FONT
    ws8.cell(row=r, column=2, value=value).alignment = LEFT
    ws8.cell(row=r, column=1).border = BORDER
    ws8.cell(row=r, column=2).border = BORDER
    ws8.row_dimensions[r].height = 30
    if r % 2 == 0:
        ws8.cell(row=r, column=1).fill = BAND_FILL
        ws8.cell(row=r, column=2).fill = BAND_FILL
    r += 1
ws8.column_dimensions["A"].width = 26
ws8.column_dimensions["B"].width = 110

# ── Final Sheet: Performance Assessment ────────────────────────────────────
ws9 = wb.create_sheet("Performance Assessment")
write_title(ws9, "Performance Assessment", "Baseline load test verdict and recommendations")

ws9["A4"] = "Overall Result"
ws9["A4"].font = LABEL_FONT
ws9["B4"] = result_flag
ws9["B4"].font = Font(bold=True, size=16, color=flag_font_color)
ws9["B4"].fill = PatternFill("solid", fgColor=flag_fill)
ws9["B4"].alignment = CENTER
ws9.row_dimensions[4].height = 24

assessment_text = (
    "No — the application did NOT successfully sustain 100 concurrent users. Only "
    f"{overall['successful_requests']} of {overall['total_requests']} requests ({overall['success_rate_pct']}%) "
    "completed successfully; the login endpoint alone timed out or errored on the large majority of attempts, and "
    "every downstream authenticated request from a VU that never obtained a session cookie failed with 401. "
    "Backend CPU was observed pegged at ~99% (single core) during login bursts, confirming the backend was fully "
    "CPU-saturated rather than network- or database-bound."
)
ws9["A6"] = "Assessment"
ws9["A6"].font = SECTION_FONT
ws9.merge_cells("A7:H7")
ws9["A7"] = assessment_text
ws9["A7"].alignment = LEFT
ws9.row_dimensions[7].height = 75

ws9["A9"] = "Root Cause"
ws9["A9"].font = SECTION_FONT
root_cause = (
    "api/auth/service.py::login_user() calls api.core.security.verify_password() — a synchronous, CPU-bound "
    "bcrypt check (cost factor 12) — directly inside an `async def` route handler, with no asyncio.to_thread()/"
    "run_in_executor() offload. FastAPI's single uvicorn worker runs one asyncio event loop; each bcrypt call "
    "blocks that event loop for its full duration (~200-300ms+ under contention), so concurrent login requests "
    "queue up strictly serially instead of running in parallel. At 100 concurrent VUs this queue outgrew k6's "
    "60s request timeout, producing the timeouts and 401/500 cascade seen across every other endpoint."
)
ws9.merge_cells("A10:H10")
ws9["A10"] = root_cause
ws9["A10"].alignment = LEFT
ws9.row_dimensions[10].height = 75

ws9["A12"] = "Top 5 Slowest Endpoints"
ws9["A12"].font = SECTION_FONT
slow_header = 13
for c, h in enumerate(["Endpoint", "Method", "Avg Response (ms)", "P99 (ms)", "Requests"], start=1):
    ws9.cell(row=slow_header, column=c, value=h)
style_header_row(ws9, slow_header, 5)
top5 = sorted(endpoints, key=lambda e: e["avg_response_ms"], reverse=True)[:5]
r = slow_header
for e in top5:
    r += 1
    ws9.cell(row=r, column=1, value=e["endpoint"])
    ws9.cell(row=r, column=2, value=e["method"])
    ws9.cell(row=r, column=3, value=e["avg_response_ms"]).number_format = "#,##0.00"
    ws9.cell(row=r, column=4, value=e["p99_ms"]).number_format = "#,##0.00"
    ws9.cell(row=r, column=5, value=e["requests"]).number_format = "#,##0"
band_rows(ws9, slow_header + 1, r, 5)
response_time_conditional_format(ws9, f"C{slow_header+1}:C{r}")

rec_start = r + 3
ws9.cell(row=rec_start, column=1, value="Top Recommendations for Optimization").font = SECTION_FONT
recommendations = [
    "1. Offload bcrypt verify/hash calls in api/core/security.py to a thread pool (asyncio.to_thread) or a "
    "process pool, so password hashing no longer blocks the event loop — this is the single highest-impact fix "
    "given the root cause above.",
    "2. Run uvicorn with multiple workers (--workers N, N = CPU cores) or behind gunicorn's UvicornWorker, so a "
    "slow synchronous call in one worker doesn't stall all in-flight requests.",
    "3. Consider a cheaper bcrypt cost factor for non-production tiers, or a faster KDF (argon2id tuned for "
    "target latency) if login throughput remains a bottleneck after (1) and (2).",
    "4. Add a per-endpoint timeout/circuit breaker so a saturated login path degrades gracefully (fast 503) "
    "instead of holding connections open until the client's 60s timeout.",
    "5. Re-run this baseline after the above fixes, and add a soak/spike test alongside it — this run only "
    "establishes the current (pre-optimization) ceiling, not sustained-load behavior.",
]
r = rec_start
for rec in recommendations:
    r += 1
    ws9.merge_cells(f"A{r}:H{r}")
    ws9.cell(row=r, column=1, value=rec).alignment = LEFT
    ws9.row_dimensions[r].height = 32

readiness_row = r + 2
ws9.cell(row=readiness_row, column=1, value="Overall Production Readiness").font = SECTION_FONT
ws9.merge_cells(f"A{readiness_row+1}:H{readiness_row+1}")
ws9.cell(
    row=readiness_row + 1,
    column=1,
    value=(
        "NOT production-ready at 100 concurrent users in its current form. The auth path has a clear, well-"
        "understood, and fixable bottleneck (blocking bcrypt on the event loop) rather than a fundamental "
        "architecture problem — Postgres and Redis stayed idle throughout (see Resource Usage) and the "
        "non-auth endpoints returned in low single-digit milliseconds whenever they could execute. Recommend "
        "applying fix #1 and #2 above and re-running this exact baseline before considering this endpoint "
        "production-ready."
    ),
).alignment = LEFT
ws9.row_dimensions[readiness_row + 1].height = 60

ws9.column_dimensions["A"].width = 16
for col in "BCDEFGH":
    ws9.column_dimensions[col].width = 14

# ── Reorder & save ─────────────────────────────────────────────────────────
order = [
    "Executive Summary", "Overall Performance", "Endpoint Performance", "HTTP Status Codes",
    "Resource Usage", "Response Time Distribution", "Raw Request Log", "Test Configuration",
    "Performance Assessment",
]
wb._sheets = [wb[name] for name in order]
for ws_ in wb.worksheets:
    ws_.sheet_view.showGridLines = False

wb.save(OUT_XLSX)
print("Saved", OUT_XLSX)
