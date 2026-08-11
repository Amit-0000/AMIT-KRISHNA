"""Builds the Selenium suite's Excel deliverables from
e2e/results/selenium_summary.json (parse_results.py's output):

  e2e/results/Automation_Test_Report.xlsx  (6 sheets, see below)
  e2e/results/Passed_Test_Cases.xlsx
  e2e/results/Failed_Test_Cases.xlsx
  e2e/results/Summary_Report.xlsx

Pure openpyxl, no pandas — same convention and palette as
voiceguard/performance/build_excel.py and voiceguard/reports/
build_master_test_report.py, so every Excel report in this repo reads as
one family.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).parent
RESULTS_DIR = ROOT.parent / "results"
SUMMARY_JSON = RESULTS_DIR / "selenium_summary.json"

OUT_MAIN = RESULTS_DIR / "Automation_Test_Report.xlsx"
OUT_PASSED = RESULTS_DIR / "Passed_Test_Cases.xlsx"
OUT_FAILED = RESULTS_DIR / "Failed_Test_Cases.xlsx"
OUT_SUMMARY = RESULTS_DIR / "Summary_Report.xlsx"

# ─── Palette / styles (matches performance/build_excel.py) ──────────────────
NAVY = "1F2937"
ACCENT = "2563EB"
LIGHT_BAND = "F3F4F6"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
GREEN_FONT = "1E7B34"
YELLOW = "FFEB9C"
YELLOW_FONT = "9C6500"
RED = "FFC7CE"
RED_FONT = "9C0006"
GREY = "E5E7EB"
GREY_FONT = "6B7280"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=16)
SUBTITLE_FONT = Font(italic=True, color=GREY_FONT, size=10)
SECTION_FONT = Font(bold=True, color=ACCENT, size=12)
LABEL_FONT = Font(bold=True, color=NAVY)
BAND_FILL = PatternFill("solid", fgColor=LIGHT_BAND)
PASS_FILL = PatternFill("solid", fgColor=GREEN)
FAIL_FILL = PatternFill("solid", fgColor=RED)
SKIP_FILL = PatternFill("solid", fgColor=GREY)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

PRIORITY_FONT_COLOR = {
    "critical": RED_FONT,
    "high": YELLOW_FONT,
    "medium": ACCENT,
    "low": GREY_FONT,
    "unspecified": GREY_FONT,
}


def style_header_row(ws: Worksheet, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def band_rows(ws: Worksheet, start_row: int, end_row: int, n_cols: int) -> None:
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = BAND_FILL
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = BORDER


def autosize(ws: Worksheet, widths: dict[str, int] | None = None, min_width=10, max_width=60) -> None:
    widths = widths or {}
    col_max: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            length = len(str(cell.value))
            col_max[col] = max(col_max.get(col, 0), length)
    for col, length in col_max.items():
        w = widths.get(col, max(min_width, min(max_width, length + 3)))
        ws.column_dimensions[col].width = w


def write_title(ws: Worksheet, title: str, subtitle: str = "") -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT


def status_fill(status: str) -> PatternFill:
    if status == "passed":
        return PASS_FILL
    if status == "skipped":
        return SKIP_FILL
    return FAIL_FILL


def status_font_color(status: str) -> str:
    if status == "passed":
        return GREEN_FONT
    if status == "skipped":
        return GREY_FONT
    return RED_FONT


TEST_ROW_HEADERS = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]


def write_test_rows_sheet(ws: Worksheet, rows: list[dict], header_row: int = 4) -> int:
    for c, h in enumerate(TEST_ROW_HEADERS, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(TEST_ROW_HEADERS))
    r = header_row
    for row in rows:
        r += 1
        ws.cell(row=r, column=1, value=row["test_id"])
        ws.cell(row=r, column=2, value=row["module"])
        ws.cell(row=r, column=3, value=row["test_name"])
        status_cell = ws.cell(row=r, column=4, value=row["status"])
        status_cell.fill = status_fill(row["status"])
        status_cell.font = Font(bold=True, color=status_font_color(row["status"]))
        ws.cell(row=r, column=5, value=row["duration_s"]).number_format = "#,##0.000"
        priority_cell = ws.cell(row=r, column=6, value=row["priority"])
        priority_cell.font = Font(bold=True, color=PRIORITY_FONT_COLOR.get(row["priority"], GREY_FONT))
    band_rows(ws, header_row + 1, r, len(TEST_ROW_HEADERS))
    autosize(ws, {"A": 12, "B": 24, "C": 48, "D": 12, "E": 16, "F": 12})
    ws.freeze_panes = f"A{header_row + 1}"
    return r


def build_automation_test_report(summary: dict) -> None:
    wb = Workbook()
    totals = summary["totals"]
    test_rows = summary["test_rows"]
    passed_rows = [t for t in test_rows if t["status"] == "passed"]
    failed_rows_ = [t for t in test_rows if t["status"] not in ("passed", "skipped")]
    skipped_rows = [t for t in test_rows if t["status"] == "skipped"]

    # ── Sheet 1: Executed Test Cases ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executed Test Cases"
    write_title(ws1, "VoiceGuard — Selenium Automation Test Report", "Executed Test Cases")
    write_test_rows_sheet(ws1, test_rows)

    # ── Sheet 2: Passed Tests ───────────────────────────────────────────
    ws2 = wb.create_sheet("Passed Tests")
    write_title(ws2, "Passed Tests", f"{len(passed_rows)} of {totals['executed']} executed")
    write_test_rows_sheet(ws2, passed_rows)

    # ── Sheet 3: Failed Tests ───────────────────────────────────────────
    ws3 = wb.create_sheet("Failed Tests")
    write_title(ws3, "Failed Tests", f"{len(failed_rows_)} of {totals['executed']} executed")
    write_test_rows_sheet(ws3, failed_rows_)

    # ── Sheet 4: Skipped Tests ───────────────────────────────────────────
    ws4 = wb.create_sheet("Skipped Tests")
    write_title(ws4, "Skipped Tests", f"{len(skipped_rows)} of {totals['total']} total")
    write_test_rows_sheet(ws4, skipped_rows)

    # ── Sheet 5: Execution Metrics ───────────────────────────────────────
    ws5 = wb.create_sheet("Execution Metrics")
    write_title(ws5, "Execution Metrics", f"Generated {summary['generated_at']}")
    metric_rows = [
        ("Total Test Cases", totals["total"], "#,##0"),
        ("Executed", totals["executed"], "#,##0"),
        ("Passed", totals["passed"], "#,##0"),
        ("Failed", totals["failed"], "#,##0"),
        ("Skipped", totals["skipped"], "#,##0"),
        ("Pass Rate (%)", totals["pass_pct"], "#,##0.00"),
        ("Duration (s)", summary["duration_s"], "#,##0.00"),
    ]
    ws5.cell(row=4, column=1, value="Metric")
    ws5.cell(row=4, column=2, value="Value")
    style_header_row(ws5, 4, 2)
    r = 5
    for label, value, fmt in metric_rows:
        ws5.cell(row=r, column=1, value=label)
        ws5.cell(row=r, column=2, value=value).number_format = fmt
        r += 1
    band_rows(ws5, 5, r - 1, 2)
    autosize(ws5, {"A": 26, "B": 16})

    chart = PieChart()
    chart.title = "Passed / Failed / Skipped"
    pie_header = r + 2
    ws5.cell(row=pie_header, column=1, value="Outcome")
    ws5.cell(row=pie_header, column=2, value="Count")
    for i, (label, val) in enumerate(
        [("Passed", totals["passed"]), ("Failed", totals["failed"]), ("Skipped", totals["skipped"])]
    ):
        ws5.cell(row=pie_header + 1 + i, column=1, value=label)
        ws5.cell(row=pie_header + 1 + i, column=2, value=val)
    data_ref = Reference(ws5, min_col=2, min_row=pie_header, max_row=pie_header + 3)
    cats_ref = Reference(ws5, min_col=1, min_row=pie_header + 1, max_row=pie_header + 3)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.width, chart.height = 14, 9
    ws5.add_chart(chart, f"D{4}")

    # ── Sheet 6: Defect Summary ──────────────────────────────────────────
    ws6 = wb.create_sheet("Defect Summary")
    write_title(ws6, "Defect Summary", f"{len(summary['failed_tests'])} failing test(s)")
    headers6 = ["Module", "Test", "Duration (s)", "Failure Reason"]
    hdr6 = 4
    for c, h in enumerate(headers6, start=1):
        ws6.cell(row=hdr6, column=c, value=h)
    style_header_row(ws6, hdr6, len(headers6))
    r = hdr6
    for f in summary["failed_tests"]:
        r += 1
        ws6.cell(row=r, column=1, value=f["module"])
        ws6.cell(row=r, column=2, value=f["nodeid"].split("::")[-1])
        ws6.cell(row=r, column=3, value=f["duration_s"]).number_format = "#,##0.000"
        reason_cell = ws6.cell(row=r, column=4, value=f.get("reason") or "(no detail captured)")
        reason_cell.alignment = LEFT
        ws6.row_dimensions[r].height = 45
    band_rows(ws6, hdr6 + 1, r, len(headers6))
    autosize(ws6, {"A": 22, "B": 48, "C": 14, "D": 80})

    for ws_ in wb.worksheets:
        ws_.sheet_view.showGridLines = False
    wb.save(OUT_MAIN)
    print("Saved", OUT_MAIN)


def build_standalone(rows: list[dict], title: str, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    write_title(ws, f"VoiceGuard — {title}", f"{len(rows)} test case(s)")
    write_test_rows_sheet(ws, rows)
    ws.sheet_view.showGridLines = False
    wb.save(out_path)
    print("Saved", out_path)


def build_summary_report(summary: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    totals = summary["totals"]
    write_title(ws, "VoiceGuard — Selenium Suite Summary Report", f"Generated {summary['generated_at']}")

    result_flag = "PASS" if totals["pass_pct"] >= 95 else "WARN"
    flag_fill = PASS_FILL if result_flag == "PASS" else FAIL_FILL
    flag_font = Font(bold=True, size=14, color=(GREEN_FONT if result_flag == "PASS" else RED_FONT))
    ws["A4"] = "Overall Result"
    ws["A4"].font = LABEL_FONT
    ws["B4"] = result_flag
    ws["B4"].font = flag_font
    ws["B4"].fill = flag_fill
    ws["B4"].alignment = CENTER

    rows = [
        ("Total Test Cases", totals["total"]),
        ("Executed", totals["executed"]),
        ("Passed", totals["passed"]),
        ("Failed", totals["failed"]),
        ("Skipped", totals["skipped"]),
        ("Pass Rate (%)", totals["pass_pct"]),
        ("Duration (s)", summary["duration_s"]),
    ]
    r = 6
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        r += 1
    band_rows(ws, 6, r - 1, 2)
    autosize(ws, {"A": 24, "B": 16})

    mod_header = r + 2
    ws.cell(row=mod_header, column=1, value="Module").font = LABEL_FONT
    ws.cell(row=mod_header, column=2, value="Pass %").font = LABEL_FONT
    style_header_row(ws, mod_header, 2)
    rr = mod_header
    for m in summary["module_breakdown"]:
        rr += 1
        ws.cell(row=rr, column=1, value=m["module"])
        ws.cell(row=rr, column=2, value=m["pass_pct"]).number_format = "#,##0.0"
    band_rows(ws, mod_header + 1, rr, 2)

    chart = BarChart()
    chart.title = "Pass % by Module"
    chart.y_axis.title = "Pass %"
    data_ref = Reference(ws, min_col=2, min_row=mod_header, max_row=rr)
    cats_ref = Reference(ws, min_col=1, min_row=mod_header + 1, max_row=rr)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width, chart.height = 18, 10
    ws.add_chart(chart, f"D{mod_header}")

    ws.sheet_view.showGridLines = False
    wb.save(OUT_SUMMARY)
    print("Saved", OUT_SUMMARY)


def main() -> None:
    summary = json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))
    if summary.get("status") != "OK":
        print(f"No OK summary available ({summary.get('status')}); skipping Excel generation.")
        return

    test_rows = summary["test_rows"]
    passed_rows = [t for t in test_rows if t["status"] == "passed"]
    failed_rows_ = [t for t in test_rows if t["status"] not in ("passed", "skipped")]

    build_automation_test_report(summary)
    build_standalone(passed_rows, "Passed Test Cases", OUT_PASSED)
    build_standalone(failed_rows_, "Failed Test Cases", OUT_FAILED)
    build_summary_report(summary)


if __name__ == "__main__":
    main()
